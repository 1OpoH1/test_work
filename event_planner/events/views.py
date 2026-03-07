from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import viewsets, status, filters, serializers
from rest_framework.pagination import PageNumberPagination
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from django.shortcuts import get_object_or_404
from PIL import Image
from io import BytesIO
from django.core.files.base import ContentFile
import openpyxl
from django.http import HttpResponse
from datetime import datetime
from drf_spectacular.utils import extend_schema, OpenApiParameter, OpenApiExample, inline_serializer, OpenApiResponse
from drf_spectacular.types import OpenApiTypes

from .models import Event, EventImage
from locations.models import Location
from .serializers import EventSerializer, EventImageSerializer, EventImageUploadSerializer
from .permissions import IsSuperUser
from .filters import EventFilter
from .utils import parse_datetime, create_excel_workbook, format_excel_header

class EventPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100

class EventViewSet(viewsets.ModelViewSet):
    queryset = Event.objects.all()
    serializer_class = EventSerializer
    pagination_class = EventPagination

    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter
    ]

    filterset_class = EventFilter
    search_fields = ['title', 'location__name']  # поиск по названию мероприятия или места
    ordering_fields = ['title', 'start_datetime', 'end_datetime']  # разрешаем сортировку по этим полям
    ordering = ['start_datetime']  # сортировка по умолчанию

    def get_permissions(self):
        """
        Только суперпользователь может создавать/изменять/удалять.
        Для чтения доступ открыт всем, но queryset фильтруется.
        """
        if self.action in ['create', 'update', 'partial_update', 'destroy', 'upload_image']:
            permission_classes = [IsAuthenticated, IsSuperUser]
        else:
            permission_classes = [AllowAny]
        return [permission() for permission in permission_classes]

    def get_queryset(self):
        """
        Суперпользователь видит все мероприятия.
        Остальные видят только опубликованные с датой публикации не позже текущего момента.
        """
        user = self.request.user
        if user.is_authenticated and user.is_superuser:
            return Event.objects.all()
        now = timezone.now()
        return Event.objects.filter(status='published', pub_datetime__lte=now)

    def perform_create(self, serializer):
        """Автором автоматически становится текущий пользователь (суперпользователь)."""
        serializer.save(author=self.request.user)

    @extend_schema(
        request=EventImageUploadSerializer,
        responses={
            201: OpenApiResponse(
                response=EventImageSerializer,
                description='Изображение успешно загружено'
            ),
            400: OpenApiResponse(
                response=OpenApiTypes.OBJECT,
                description='Ошибка валидации'
            ),
            401: OpenApiResponse(
                response=OpenApiTypes.OBJECT,
                description='Не авторизован'
            ),
            403: OpenApiResponse(
                response=OpenApiTypes.OBJECT,
                description='Доступ запрещен (только суперпользователь)'
            ),
            404: OpenApiResponse(
                response=OpenApiTypes.OBJECT,
                description='Мероприятие не найдено'
            )
        },
        parameters=[
            OpenApiParameter(
                name='id',
                type=OpenApiTypes.INT,
                location=OpenApiParameter.PATH,
                description='ID мероприятия',
                required=True
            )
        ],
        description='Загрузка изображения для мероприятия. Первое загруженное изображение автоматически становится превью.',
        methods=['POST']
    )
    @action(detail=True, methods=['post'], url_path='upload-image')
    def upload_image(self, request, pk=None):
        """Загрузка изображения для конкретного мероприятия."""
        event = self.get_object()
        file = request.FILES.get('image')
        if not file:
            return Response({'error': 'No image provided'}, status=status.HTTP_400_BAD_REQUEST)

        image = EventImage.objects.create(event=event, image=file)

        # Если у мероприятия ещё нет превью — генерируем его из первого загруженного изображения
        if not event.preview:
            self._generate_preview(event, file)

        serializer = EventImageSerializer(image)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def _generate_preview(self, event, image_file):
        """
        Создаёт превью 200px по наименьшей стороне и сохраняет в поле preview модели Event.
        """
        try:
            img = Image.open(image_file)
            # Определяем наименьшую сторону
            min_side = min(img.width, img.height)
            scale = 200 / min_side
            new_width = int(img.width * scale)
            new_height = int(img.height * scale)
            img = img.resize((new_width, new_height), Image.LANCZOS)

            # Сохраняем в буфер
            buffer = BytesIO()
            format = 'JPEG' if img.format == 'JPEG' else 'PNG'
            img.save(buffer, format=format)
            buffer.seek(0)

            # Сохраняем в поле preview
            filename = f"preview_{event.id}.{format.lower()}"
            event.preview.save(filename, ContentFile(buffer.read()), save=True)
        except Exception as e:
            # Логирование ошибки, но не прерываем загрузку изображения
            print(f"Error generating preview: {e}")
    
    @extend_schema(
        request=inline_serializer(
            name='EventImportSerializer',
            fields={
                'file': serializers.FileField(help_text='XLSX файл с мероприятиями')
            }
        ),
        responses={
            201: inline_serializer(
                name='ImportSuccessSerializer',
                fields={
                    'created': serializers.IntegerField(),
                    'errors': serializers.ListField(child=serializers.CharField(), required=False)
                }
            ),
            400: OpenApiTypes.OBJECT,
            207: OpenApiTypes.OBJECT
        },
        description='Импорт мероприятий из XLSX-файла',
        methods=['POST']
    )
    @action(detail=False, methods=['post'], url_path='import', permission_classes=[IsAuthenticated, IsSuperUser])
    def import_events(self, request):
        """
        Импорт мероприятий из XLSX-файла.
        Формат колонок (порядок важен):
        1. Название
        2. Описание
        3. Дата и время публикации
        4. Дата и время начала проведения
        5. Дата и время завершения проведения
        6. Название места проведения
        7. Широта
        8. Долгота
        9. Рейтинг (0-25)
        """
        file = request.FILES.get('file')
        if not file:
            return Response(
                {'error': 'Необходимо загрузить XLSX-файл'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active

            created_count = 0
            errors = []
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                try:
                    if not any(row):
                        continue

                    if len(row) < 9:
                        errors.append(f"Строка {row_num} имеет недостаточно данных")
                        continue

                    title, description, pub_dt, start_dt, end_dt, location_name, lat, lon, rating = row[:9]
                    if not title or not description:
                        errors.append(f"Строка {row_num}: название и описание обязательны")
                        continue
                    
                    # Парсинг дат
                    pub_datetime = parse_datetime(pub_dt)
                    start_datetime = parse_datetime(start_dt)
                    end_datetime = parse_datetime(end_dt)
                    
                    if not all([pub_datetime, start_datetime, end_datetime]):
                        errors.append(f"Строка {row_num}: некорректный формат даты/времени")
                        continue
                    
                    # Валидация рейтинга
                    try:
                        rating = int(rating) if rating is not None else 0
                        if not (0 <= rating <= 25):
                            errors.append(f"Строка {row_num}: рейтинг должен быть от 0 до 25")
                            continue
                    except (TypeError, ValueError):
                        errors.append(f"Строка {row_num}: некорректный рейтинг")
                        continue
                    
                    # Поиск или создание места проведения
                    location = None
                    if location_name and lat is not None and lon is not None:
                        try:
                            # Пытаемся найти существующее место
                            location = Location.objects.filter(
                                name__iexact=location_name.strip()
                            ).first()
                            
                            # Если не найдено, создаём новое
                            if not location:
                                location = Location.objects.create(
                                    name=location_name.strip(),
                                    latitude=float(lat),
                                    longitude=float(lon)
                                )
                        except (ValueError, TypeError) as e:
                            errors.append(f"Строка {row_num}: ошибка координат - {str(e)}")
                            continue
                    
                    # Создание мероприятия
                    event = Event.objects.create(
                        title=title.strip(),
                        description=description.strip(),
                        pub_datetime=pub_datetime,
                        start_datetime=start_datetime,
                        end_datetime=end_datetime,
                        location=location,
                        rating=rating,
                        status='draft',  # По умолчанию черновик
                        author=request.user
                    )
                    
                    created_count += 1
                    
                except Exception as e:
                    errors.append(f"Строка {row_num}: {str(e)}")
            
            response_data = {
                'created': created_count,
                'errors': errors
            }
            
            if errors:
                response_data['message'] = f"Импортировано {created_count} мероприятий с ошибками"
                return Response(response_data, status=status.HTTP_207_MULTI_STATUS)
            else:
                return Response(response_data, status=status.HTTP_201_CREATED)
                
        except Exception as e:
            return Response(
                {'error': f'Ошибка обработки файла: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
    @action(detail=False, methods=['get'], url_path='export')
    def export_events(self, request):
        """
        Экспорт мероприятий в XLSX с фильтрацией.
        Поддерживает те же параметры фильтрации, что и список мероприятий.
        """
        # Получаем отфильтрованный queryset (используя те же фильтры)
        queryset = self.filter_queryset(self.get_queryset())
        
        # Создаём Excel-книгу
        wb, ws = create_excel_workbook()
        
        # Заголовки
        headers = [
            'ID', 'Название', 'Описание', 'Дата публикации',
            'Дата начала', 'Дата завершения', 'Место проведения',
            'Широта', 'Долгота', 'Рейтинг', 'Статус', 'Автор'
        ]
        format_excel_header(ws, headers)
        
        # Заполнение данных
        for row_num, event in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=event.id)
            ws.cell(row=row_num, column=2, value=event.title)
            ws.cell(row=row_num, column=3, value=event.description)
            ws.cell(row=row_num, column=4, value=event.pub_datetime.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=5, value=event.start_datetime.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=6, value=event.end_datetime.strftime('%Y-%m-%d %H:%M:%S'))
            
            # Информация о месте
            if event.location:
                ws.cell(row=row_num, column=7, value=event.location.name)
                ws.cell(row=row_num, column=8, value=event.location.latitude)
                ws.cell(row=row_num, column=9, value=event.location.longitude)
            
            ws.cell(row=row_num, column=10, value=event.rating)
            ws.cell(row=row_num, column=11, value=event.get_status_display())
            ws.cell(row=row_num, column=12, value=event.author.username)
        
        # Настройка HTTP-ответа
        filename = f"events_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Сохраняем книгу в ответ
        wb.save(response)
        return response