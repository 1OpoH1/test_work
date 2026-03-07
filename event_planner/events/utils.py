import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.utils import timezone
from datetime import datetime

def create_excel_workbook():
    """Создание новой рабочей книги с настройками"""
    wb = openpyxl.Workbook()
    ws = wb.active
    return wb, ws

def format_excel_header(ws, headers):
    """Форматирование заголовков в Excel"""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
        # Автоподбор ширины колонки
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = len(header) + 5

def parse_datetime(value):
    """Парсинг даты/времени из строки Excel"""
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        # Пробуем разные форматы
        formats = [
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%dT%H:%M:%S',
            '%Y-%m-%d %H:%M',
            '%d.%m.%Y %H:%M',
            '%Y-%m-%d'
        ]
        for fmt in formats:
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
    return None